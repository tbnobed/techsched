from flask import render_template, redirect, url_for, flash, request, jsonify, send_file, make_response
from flask_login import login_user, logout_user, login_required, current_user
from app import app, db, is_mobile_device
from models import User, Schedule, QuickLink, Location, EmailSettings, TicketCategory, Ticket, TicketComment, TicketHistory, TicketStatus
from forms import (
    LoginForm, RegistrationForm, ScheduleForm, AdminUserForm, EditUserForm, 
    ChangePasswordForm, QuickLinkForm, LocationForm, EmailSettingsForm
)
from datetime import datetime, timedelta, time
import pytz
import csv
import random
import string
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import json
import os
from werkzeug.utils import secure_filename
from email_utils import send_schedule_notification
from flask import session

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('calendar'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    app.logger.debug(f"Login attempt - Method: {request.method}")
    app.logger.debug(f"Session before login: {session}")

    if current_user.is_authenticated:
        app.logger.debug(f"Already authenticated user: {current_user.username}")
        app.logger.debug(f"Current session: {session}")
        return redirect(url_for('calendar'))

    form = LoginForm()
    if form.validate_on_submit():
        app.logger.debug(f"Login form submitted for email: {form.email.data}")
        user = User.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            # Set session to be permanent (7 days)
            session.permanent = True
            login_user(user, remember=form.remember_me.data)
            app.logger.info(f"User {user.username} logged in successfully")
            app.logger.debug(f"Session after login: {session}")
            next_page = request.args.get('next')
            app.logger.debug(f"Redirecting to: {next_page if next_page else 'calendar'}")
            return redirect(next_page if next_page else url_for('calendar'))
        app.logger.warning(f"Failed login attempt for email: {form.email.data}")
        flash('Invalid email or password')
    return render_template('login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('calendar'))

    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please log in.')
        return redirect(url_for('login'))
    return render_template('register.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/api/active_users')
@login_required
def get_active_users():
    """Get users who have schedules active at the current time"""
    if not current_user.is_authenticated:
        return jsonify({'error': 'Authentication required'}), 401

    try:
        # Get current time in UTC since our database stores times in UTC
        current_time = datetime.now(pytz.UTC)
        app.logger.debug(f"Current time (UTC): {current_time}")
        
        # Find who is actively scheduled right now
        active_users_query = (db.session.query(
                User, Schedule, Location
            )
            .join(Schedule, User.id == Schedule.technician_id)
            .outerjoin(Location, Schedule.location_id == Location.id)
            .filter(
                Schedule.start_time <= current_time,
                Schedule.end_time > current_time,
                ~Schedule.time_off  # Exclude time off entries
            ))
            
        app.logger.debug(f"Active users query SQL: {str(active_users_query)}")
        active_users_data = active_users_query.all()
        app.logger.debug(f"Found {len(active_users_data)} active schedules")

        # Build the result - only show currently active users
        result = []
        
        # Process active users data
        user_tz = current_user.get_timezone()
        for user, schedule, location in active_users_data:
            # Skip if any key component is None
            if not user or not schedule:
                app.logger.warning(f"Skipping schedule with missing user or schedule data")
                continue
                
            try:
                # Convert schedule times to user's timezone
                start_time = schedule.start_time.astimezone(user_tz)
                end_time = schedule.end_time.astimezone(user_tz)
                
                # Add this active user to the result
                result.append({
                    'username': user.username,
                    'color': user.color,
                    'schedule': {
                        'start_time': start_time.strftime('%H:%M'),
                        'end_time': end_time.strftime('%H:%M'),
                        'description': schedule.description or ''
                    },
                    'location': {
                        'name': location.name if location else 'No Location',
                        'description': location.description if location else ''
                    }
                })
            except Exception as inner_e:
                app.logger.error(f"Error processing schedule for user {user.id}: {str(inner_e)}")
                # Continue processing other users

        app.logger.debug(f"Returning {len(result)} active users")
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Error in get_active_users: {str(e)}")
        app.logger.exception("Detailed exception information:")
        return jsonify([]), 200  # Return empty array instead of error

@app.route('/profile')
@login_required
def profile():
    form = EditUserForm(obj=current_user)
    password_form = ChangePasswordForm()
    
    # For debugging
    print(f"is_mobile_device() in profile: {is_mobile_device()}")
    is_mobile = is_mobile_device()  # Force evaluation
    print(f"is_mobile value in profile: {is_mobile}")
    
    if is_mobile_device():
        # Use mobile template with timezone list
        return render_template('mobile_profile.html', 
                             form=form, 
                             password_form=password_form,
                             timezones=pytz.common_timezones)
    
    # Use desktop template
    return render_template('profile.html', form=form, password_form=password_form)

@app.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    color = request.form.get('color')
    if color:
        try:
            current_user.color = color
            db.session.commit()
            flash('Profile updated successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating profile: {str(e)}")
            flash('Error updating profile. Please try again.')
    return redirect(url_for('profile'))

@app.route('/profile/change-password', methods=['POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            current_user.set_password(form.new_password.data)
            db.session.commit()
            flash('Password updated successfully!')
            return redirect(url_for('profile'))
        flash('Current password is incorrect')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}')
    return redirect(url_for('profile'))


@app.route('/toggle-theme', methods=['POST'])
@login_required
def toggle_theme():
    try:
        # Toggle between light and dark theme
        current_theme = current_user.theme_preference or 'dark'  # Default to dark if None
        new_theme = 'light' if current_theme == 'dark' else 'dark'
        
        # Update user preference
        current_user.theme_preference = new_theme
        db.session.commit()
        
        flash(f'Theme updated to {new_theme} mode')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error toggling theme: {str(e)}")
        flash('Error updating theme preference. Please try again.')
    
    # Redirect back to the page they came from or default to profile
    return redirect(request.referrer or url_for('profile'))


@app.route('/admin/locations', methods=['GET', 'POST'])
@login_required
def admin_locations():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    form = LocationForm()
    if form.validate_on_submit():
        try:
            location = Location(
                name=form.name.data,
                description=form.description.data,
                active=form.active.data
            )
            db.session.add(location)
            db.session.commit()
            flash('Location added successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error creating location: {str(e)}")
            flash('Error creating location. Please try again.')

    locations = Location.query.order_by(Location.name).all()
    return render_template('admin/locations.html', locations=locations, form=form)

@app.route('/calendar')
@login_required
def calendar():
    week_start = request.args.get('week_start')
    location_filter = request.args.get('location_id', type=int)

    if week_start:
        week_start = datetime.strptime(week_start, '%Y-%m-%d')
        week_start = current_user.get_timezone().localize(
            week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        )
    else:
        week_start = datetime.now(current_user.get_timezone())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start -= timedelta(days=week_start.weekday())

    # Convert to UTC for database query
    week_start_utc = week_start.astimezone(pytz.UTC)
    week_end_utc = (week_start + timedelta(days=7)).astimezone(pytz.UTC)

    # Query schedules in UTC with optional location filter
    query = Schedule.query.filter(
        Schedule.start_time >= week_start_utc,
        Schedule.start_time < week_end_utc
    )

    if location_filter:
        query = query.filter(Schedule.location_id == location_filter)

    schedules = query.all()

    # Convert schedule times to user's timezone
    user_tz = current_user.get_timezone()
    for schedule in schedules:
        if schedule.start_time.tzinfo is None:
            schedule.start_time = pytz.UTC.localize(schedule.start_time)
        if schedule.end_time.tzinfo is None:
            schedule.end_time = pytz.UTC.localize(schedule.end_time)

        schedule.start_time = schedule.start_time.astimezone(user_tz)
        schedule.end_time = schedule.end_time.astimezone(user_tz)

    form = ScheduleForm()
    if current_user.is_admin:
        form.technician.choices = [(u.id, u.username) for u in User.query.all()]
    else:
        form.technician.choices = [(current_user.id, current_user.username)]
        form.technician.data = current_user.id

    # Add location choices to the form
    form.location_id.choices = [(l.id, l.name) for l in Location.query.filter_by(active=True).order_by(Location.name).all()]

    # Get all active locations for the filter dropdown
    locations = Location.query.filter_by(active=True).order_by(Location.name).all()

    # Debug the mobile detection
    print(f"is_mobile_device() in calendar: {is_mobile_device()}")
    is_mobile = is_mobile_device()  # Force evaluation
    print(f"is_mobile value: {is_mobile}")
    
    # Check if user is on a mobile device
    if is_mobile_device():
        print("Using mobile template for calendar")
        # Get active tickets for the sidebar
        from ticket_routes import get_active_sidebar_tickets
        active_sidebar_tickets = get_active_sidebar_tickets()
        
        return render_template('mobile_calendar.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            locations=locations,
                            selected_location=location_filter,
                            today=datetime.now(current_user.get_timezone()),
                            datetime=datetime,
                            timedelta=timedelta,
                            active_sidebar_tickets=active_sidebar_tickets)
    else:
        return render_template('calendar.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            locations=locations,
                            selected_location=location_filter,
                            today=datetime.now(current_user.get_timezone()),
                            datetime=datetime,
                            timedelta=timedelta)

@app.route('/schedule/new', methods=['GET', 'POST'])
@login_required
def new_schedule():
    # Get the week_start parameter to maintain the same view
    week_start = request.args.get('week_start') or request.form.get('week_start')
    personal_view = request.args.get('personal_view') == 'true' or request.form.get('personal_view') == 'true'
    
    form = ScheduleForm()

    # Set up technician choices
    if current_user.is_admin:
        form.technician.choices = [(u.id, u.username) for u in User.query.all()]
    else:
        form.technician.choices = [(current_user.id, current_user.username)]
        form.technician.data = current_user.id

    # Set up location choices
    locations = Location.query.filter_by(active=True).order_by(Location.name).all()
    form.location_id.choices = [(l.id, l.name) for l in locations]
    # Add an empty choice if no locations exist
    if not locations:
        form.location_id.choices = [(0, 'No locations available')]

    # Check if we're dealing with a mobile form submission (with schedule_date, start_hour, end_hour)
    schedule_date = request.form.get('schedule_date')
    start_hour = request.form.get('start_hour')
    end_hour = request.form.get('end_hour')
    
    # Debug log to see what's being submitted
    app.logger.debug(f"Form submission: {request.form}")
    
    is_mobile_submission = schedule_date and start_hour and end_hour
    
    if is_mobile_submission:
        # We're getting data from the mobile form
        app.logger.debug(f"Mobile form data received: date={schedule_date}, start={start_hour}, end={end_hour}")
        try:
            # Parse the date
            date_obj = datetime.strptime(schedule_date, '%Y-%m-%d').date()
            
            # Create start_time combining date and start_hour
            start_hour_int = int(start_hour)
            start_time_obj = datetime.combine(date_obj, time(hour=start_hour_int, minute=0))
            
            # Create end_time combining date and end_hour
            end_hour_int = int(end_hour)
            # If end hour is 0 (midnight), it should be the next day
            if end_hour_int == 0:
                # Set to midnight of the same day, will add a day later for UTC
                end_time_obj = datetime.combine(date_obj, time(hour=0, minute=0))
            else:
                end_time_obj = datetime.combine(date_obj, time(hour=end_hour_int, minute=0))
            
            # Assign to form
            form.start_time.data = start_time_obj
            form.end_time.data = end_time_obj
            
            # Handle form fields
            form.description.data = request.form.get('description', '')
            
            # Handle time_off checkbox
            time_off_val = request.form.get('time_off')
            form.time_off.data = bool(time_off_val == 'on' or time_off_val == 'true' or time_off_val == '1')
            
            # Get the location_id if present
            location_id = request.form.get('location_id')
            if location_id and location_id.isdigit():
                form.location_id.data = int(location_id)
            else:
                form.location_id.data = 0
                
            # Mobile validation successful
            is_mobile_validation_successful = True
            
        except Exception as e:
            app.logger.error(f"Error parsing mobile form data: {str(e)}")
            flash('Invalid date or time format. Please try again.')
            return redirect(url_for('calendar', week_start=week_start))

    if form.validate_on_submit() or is_mobile_submission:
        try:
            app.logger.debug(f"Processing form data: {request.form}")
            schedule_id = request.form.get('schedule_id')
            technician_id = form.technician.data if current_user.is_admin else current_user.id
            
            # Also check if we have a week_start in the form
            if not week_start:
                week_start = request.form.get('week_start')

            user_tz = current_user.get_timezone()
            start_time = user_tz.localize(form.start_time.data)
            end_time = user_tz.localize(form.end_time.data)

            start_time_utc = start_time.astimezone(pytz.UTC)
            end_time_utc = end_time.astimezone(pytz.UTC)

            if end_time.hour == 0 and end_time.minute == 0:
                end_time_utc = end_time_utc + timedelta(days=1)

            if end_time_utc <= start_time_utc:
                flash('End time must be after start time.')
                return redirect(url_for('calendar', week_start=week_start))

            overlapping_query = Schedule.query.filter(
                Schedule.technician_id == technician_id,
                Schedule.id != (int(schedule_id) if schedule_id else None),
                Schedule.start_time < end_time_utc,
                Schedule.end_time > start_time_utc
            )

            overlapping_schedules = overlapping_query.first()

            if overlapping_schedules and not form.time_off.data:
                flash('Schedule conflicts with existing appointments.')
                if personal_view:
                    return redirect(url_for('personal_schedule', week_start=week_start))
                else:
                    return redirect(url_for('calendar', week_start=week_start))

            # Check if we have repeat days selected
            repeat_days = form.repeat_days.data
            
            if schedule_id:
                # Editing an existing schedule - doesn't support multi-day editing
                schedule = Schedule.query.get_or_404(schedule_id)
                if schedule.technician_id != current_user.id and not current_user.is_admin:
                    flash('You do not have permission to edit this schedule.')
                    if personal_view:
                        return redirect(url_for('personal_schedule', week_start=week_start))
                    else:
                        return redirect(url_for('calendar', week_start=week_start))

                old_desc = schedule.description
                schedule.start_time = start_time_utc
                schedule.end_time = end_time_utc
                schedule.description = form.description.data
                schedule.time_off = form.time_off.data
                schedule.location_id = form.location_id.data if form.location_id.data != 0 else None
                if current_user.is_admin:
                    schedule.technician_id = technician_id
                send_schedule_notification(schedule, 'updated', f"Schedule updated by {current_user.username}")
                
                db.session.commit()
                flash('Schedule updated successfully!')
                
            else:
                # Creating new schedule(s)
                schedules_created = 0
                
                if repeat_days:
                    # Multi-day scheduling
                    dates = repeat_days.split(',')
                    app.logger.debug(f"Creating schedules for multiple days: {dates}")
                    
                    if not dates:
                        flash('No valid dates selected for scheduling.')
                        if personal_view:
                            return redirect(url_for('personal_schedule', week_start=week_start))
                        else:
                            return redirect(url_for('calendar', week_start=week_start))
                    
                    # Create a schedule for each selected day
                    for date_str in dates:
                        # Parse the date
                        day_date = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
                        
                        # Create a new datetime using the date from day_date and time from start_time/end_time
                        day_start_time = user_tz.localize(
                            datetime.combine(day_date, datetime.min.time().replace(
                                hour=start_time.hour, minute=start_time.minute
                            ))
                        )
                        day_end_time = user_tz.localize(
                            datetime.combine(day_date, datetime.min.time().replace(
                                hour=end_time.hour, minute=end_time.minute
                            ))
                        )
                        
                        # Convert to UTC
                        day_start_time_utc = day_start_time.astimezone(pytz.UTC)
                        day_end_time_utc = day_end_time.astimezone(pytz.UTC)
                        
                        # Handle midnight end time
                        if end_time.hour == 0 and end_time.minute == 0:
                            day_end_time_utc = day_end_time_utc + timedelta(days=1)
                            
                        # Check for overlapping schedules for this specific day
                        overlapping_query = Schedule.query.filter(
                            Schedule.technician_id == technician_id,
                            Schedule.start_time < day_end_time_utc,
                            Schedule.end_time > day_start_time_utc
                        )
                        overlapping_schedule = overlapping_query.first()
                        
                        if overlapping_schedule and not form.time_off.data:
                            app.logger.warning(f"Skipping schedule for {date_str} due to conflict")
                            continue
                            
                        # Create the schedule for this day
                        schedule = Schedule(
                            technician_id=technician_id,
                            start_time=day_start_time_utc,
                            end_time=day_end_time_utc,
                            description=form.description.data,
                            time_off=form.time_off.data,
                            location_id=form.location_id.data if form.location_id.data != 0 else None
                        )
                        db.session.add(schedule)
                        schedules_created += 1
                        
                    if schedules_created > 0:
                        send_schedule_notification(schedule, 'created', 
                            f"Multiple schedules created by {current_user.username}")
                        db.session.commit()
                        flash(f'{schedules_created} schedules created successfully!')
                    else:
                        flash('No schedules could be created due to conflicts with existing schedules.')
                        
                else:
                    # Single day scheduling
                    schedule = Schedule(
                        technician_id=technician_id,
                        start_time=start_time_utc,
                        end_time=end_time_utc,
                        description=form.description.data,
                        time_off=form.time_off.data,
                        location_id=form.location_id.data if form.location_id.data != 0 else None
                    )
                    db.session.add(schedule)
                    send_schedule_notification(schedule, 'created', f"Schedule created by {current_user.username}")
                    db.session.commit()
                    flash('Schedule created successfully!')
            if personal_view:
                return redirect(url_for('personal_schedule', week_start=week_start))
            else:
                return redirect(url_for('calendar', week_start=week_start))

        except Exception as e:
            db.session.rollback()
            flash('Error saving schedule. Please check the time entries.')
            app.logger.error(f"Error saving schedule: {str(e)}")
            if personal_view:
                return redirect(url_for('personal_schedule', week_start=week_start))
            else:
                return redirect(url_for('calendar', week_start=week_start))

    if personal_view:
        return redirect(url_for('personal_schedule', week_start=week_start))
    else:
        return redirect(url_for('calendar', week_start=week_start))

@app.route('/schedule/delete/<int:schedule_id>')
@login_required
def delete_schedule(schedule_id):
    # Get current week start to maintain the same view
    week_start = request.args.get('week_start') or request.form.get('week_start')
    personal_view = request.args.get('personal_view') == 'true' or request.form.get('personal_view') == 'true'
    
    schedule = Schedule.query.get_or_404(schedule_id)

    if schedule.technician_id != current_user.id and not current_user.is_admin:
        flash('You do not have permission to delete this schedule.')
        if personal_view:
            return redirect(url_for('personal_schedule', week_start=week_start))
        else:
            return redirect(url_for('calendar', week_start=week_start))

    try:
        send_schedule_notification(schedule, 'deleted', f"Schedule deleted by {current_user.username}")
        db.session.delete(schedule)
        db.session.commit()
        flash('Schedule deleted successfully!')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting schedule.')
        app.logger.error(f"Error deleting schedule: {str(e)}")

    # Redirect back to the same week view
    if personal_view:
        return redirect(url_for('personal_schedule', week_start=week_start))
    else:
        return redirect(url_for('calendar', week_start=week_start))

@app.route('/schedule/copy_previous_week', methods=['POST'])
@login_required
def copy_previous_week_schedules():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        if not request.form.get('csrf_token') or not request.form.get('csrf_token') == request.form.get('csrf_token'):
            flash('Invalid request. Please try again.')
            return redirect(url_for('calendar'))

        target_week_start_str = request.form.get('target_week_start')
        if not target_week_start_str:
            flash('Invalid week start date.')
            return redirect(url_for('calendar'))

        user_tz = current_user.get_timezone()

        # Convert target week start to user's timezone
        target_week_start = user_tz.localize(
            datetime.strptime(target_week_start_str, '%Y-%m-%d')
        )
        previous_week_start = target_week_start - timedelta(days=7)

        # Convert to UTC for database operations
        target_week_start_utc = target_week_start.astimezone(pytz.UTC)
        target_week_end_utc = (target_week_start + timedelta(days=7)).astimezone(pytz.UTC)
        previous_week_start_utc = previous_week_start.astimezone(pytz.UTC)
        previous_week_end_utc = (previous_week_start + timedelta(days=7)).astimezone(pytz.UTC)

        app.logger.debug(f"Copying schedules from {previous_week_start_utc} to {previous_week_end_utc}")
        app.logger.debug(f"To target week: {target_week_start_utc} to {target_week_end_utc}")

        # Get previous week's schedules
        previous_schedules = Schedule.query.filter(
            Schedule.start_time >= previous_week_start_utc,
            Schedule.start_time < previous_week_end_utc
        ).all()

        if not previous_schedules:
            flash('No schedules found in previous week to copy.')
            return redirect(url_for('calendar'))

        try:
            # Delete existing schedules in target week
            Schedule.query.filter(
                Schedule.start_time >= target_week_start_utc,
                Schedule.start_time < target_week_end_utc
            ).delete()

            # Copy schedules to new week
            time_difference = target_week_start_utc - previous_week_start_utc

            for schedule in previous_schedules:
                new_schedule = Schedule(
                    technician_id=schedule.technician_id,
                    start_time=schedule.start_time + time_difference,
                    end_time=schedule.end_time + time_difference,
                    description=schedule.description,
                    time_off=schedule.time_off,
                    location_id=schedule.location_id
                )
                db.session.add(new_schedule)

            db.session.commit()
            flash(f'Successfully copied {len(previous_schedules)} schedules from previous week!')

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Database error in copy_previous_week_schedules: {str(e)}")
            flash('Error copying schedules from previous week.')

    except Exception as e:
        app.logger.error(f"Error in copy_previous_week_schedules: {str(e)}")
        flash('Error copying schedules from previous week.')

    return redirect(url_for('calendar', week_start=target_week_start_str))

@app.route('/update_timezone', methods=['POST'])
@login_required
def update_timezone():
    timezone = request.form.get('timezone')
    if timezone in pytz.all_timezones:
        current_user.timezone = timezone
        db.session.commit()
        flash('Timezone updated successfully!')
    else:
        flash('Invalid timezone')
    return redirect(request.referrer or url_for('calendar'))

@app.route('/admin/dashboard', methods=['GET', 'POST'])
@login_required
def admin_dashboard():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    users = User.query.all()
    quick_links = QuickLink.query.order_by(QuickLink.order.asc()).all()
    form = AdminUserForm()
    edit_form = EditUserForm()
    return render_template('admin/dashboard.html', 
                         users=users, 
                         form=form, 
                         edit_form=edit_form,
                         quick_links=quick_links)

@app.route('/admin/create_user', methods=['POST'])
@login_required
def admin_create_user():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    form = AdminUserForm()
    if form.validate_on_submit():
        try:
            # Convert email to lowercase for case-insensitive searching
            email = form.email.data.lower() if form.email.data else ""
            username = form.username.data
            
            # Check if user already exists (case-insensitive using PostgreSQL LOWER)
            existing_email_user = User.query.filter(db.func.lower(User.email) == db.func.lower(email)).first()
            if existing_email_user:
                flash('Email already registered.')
                return redirect(url_for('admin_dashboard'))
                
            # Check if username already exists (case-insensitive using PostgreSQL LOWER)
            if username:
                existing_user = User.query.filter(db.func.lower(User.username) == db.func.lower(username)).first()
                if existing_user:
                    flash('Username already registered. Please choose another username.')
                    return redirect(url_for('admin_dashboard'))

            # Create new user
            user = User(
                username=form.username.data,
                email=email,  # Store email in lowercase
                color=form.color.data or '#3498db',  # Default color if none provided
                is_admin=form.is_admin.data,
                timezone=form.timezone.data or 'America/Los_Angeles'  # Default timezone
            )
            user.set_password(form.password.data)

            # Log the creation attempt
            app.logger.info(f"Creating new user with username: {user.username}, email: {user.email}, timezone: {user.timezone}")

            db.session.add(user)
            db.session.commit()
            flash('User created successfully!')

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error creating user: {str(e)}")
            flash('Error creating user. Please check the form and try again.')
    else:
        # Log form validation errors
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{field}: {error}")
                app.logger.error(f"Form validation error - {field}: {error}")

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_user(user_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    user = User.query.get_or_404(user_id)
    form = EditUserForm()

    # Debug log for incoming request
    app.logger.debug(f"Edit user request for user_id {user_id}")
    app.logger.debug(f"Request method: {request.method}")
    app.logger.debug(f"Form data: {request.form}")

    if request.method == 'GET':
        form.username.data = user.username
        form.email.data = user.email
        form.color.data = user.color
        form.is_admin.data = user.is_admin
        form.timezone.data = user.timezone
        return render_template('admin/edit_user.html', 
                            user=user,
                            form=form)

    if request.method == 'POST':
        # Get form data directly from request.form
        username = request.form.get('username')
        email = request.form.get('email')
        
        # Convert email to lowercase for case-insensitivity
        if email:
            email = email.lower()
            
        color = request.form.get('color')
        password = request.form.get('password')
        timezone = request.form.get('timezone')
        is_admin = request.form.get('is_admin') == 'on'

        app.logger.debug(f"Processed form data: username={username}, email={email}, color={color}, timezone={timezone}, is_admin={is_admin}")

        try:
            # Check if username is already taken by another user (case-insensitive using PostgreSQL LOWER)
            if username:
                username_conflict = User.query.filter(
                    User.id != user_id,
                    db.func.lower(User.username) == db.func.lower(username)
                ).first()
                if username_conflict:
                    flash(f'Username "{username}" is already taken. Please use a different username.')
                    return redirect(url_for('admin_dashboard'))
            
            # Check if email is already taken by another user (case-insensitive using PostgreSQL LOWER)
            if email:
                email_conflict = User.query.filter(
                    User.id != user_id,
                    db.func.lower(User.email) == db.func.lower(email)
                ).first()
                if email_conflict:
                    flash(f'Email "{email}" is already registered to another user.')
                    return redirect(url_for('admin_dashboard'))
            
            # Update user fields
            user.username = username
            # Store email in lowercase for case-insensitive handling
            user.email = email.lower() if email else ""
            user.color = color
            user.is_admin = is_admin
            user.timezone = timezone

            if password:
                user.set_password(password)

            # Commit changes and verify
            db.session.commit()

            # Verify the changes were saved
            updated_user = User.query.get(user_id)
            app.logger.debug(f"Updated user values: username={updated_user.username}, email={updated_user.email}, color={updated_user.color}, timezone={updated_user.timezone}, is_admin={updated_user.is_admin}")

            flash('User updated successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating user: {str(e)}")
            flash('Error updating user. Please try again.')

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_user/<int:user_id>')
@login_required
def admin_delete_user(user_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    if current_user.id == user_id:
        flash('Cannot delete your own account.')
        return redirect(url_for('admin_dashboard'))

    user = User.query.get_or_404(user_id)
    
    # Prevent deletion of the System user
    if user.username == "System":
        flash('Cannot delete the System user as it is required for system operations.')
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Get a special system user to reassign content to
        # Create one if it doesn't exist
        system_user = User.query.filter_by(username="System").first()
        if not system_user:
            system_user = User(
                username="System",
                email="system@example.com",
                is_admin=False
            )
            system_user.set_password(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20)))
            db.session.add(system_user)
            db.session.flush()  # Get the ID without committing
        
        # Reassign ticket comments to system user
        from models import TicketComment
        ticket_comments = TicketComment.query.filter_by(user_id=user_id).all()
        for comment in ticket_comments:
            comment.user_id = system_user.id
        
        # Reassign ticket history entries to system user
        from models import TicketHistory
        ticket_history = TicketHistory.query.filter_by(user_id=user_id).all()
        for history in ticket_history:
            history.user_id = system_user.id
        
        # Reassign tickets created by this user
        from models import Ticket
        created_tickets = Ticket.query.filter_by(created_by=user_id).all()
        for ticket in created_tickets:
            ticket.created_by = system_user.id
        
        # Remove ticket assignments
        assigned_tickets = Ticket.query.filter_by(assigned_to=user_id).all()
        for ticket in assigned_tickets:
            ticket.assigned_to = None
            
        # Delete associated schedules
        Schedule.query.filter_by(technician_id=user_id).delete()
        
        # Now delete the user
        db.session.delete(user)
        db.session.commit()
        flash('User and associated data deleted successfully!')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting user.')
        app.logger.error(f"Error deleting user: {str(e)}")

    return redirect(url_for('admin_dashboard'))

@app.route('/personal_schedule')
@login_required
def personal_schedule():
    week_start = request.args.get('week_start')
    if week_start:
        week_start = datetime.strptime(week_start, '%Y-%m-%d')
        week_start = current_user.get_timezone().localize(
            week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        )
    else:
        week_start = datetime.now(current_user.get_timezone())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start -= timedelta(days=week_start.weekday())

    # Convert to UTC for database query
    week_start_utc = week_start.astimezone(pytz.UTC)
    week_end_utc = (week_start + timedelta(days=7)).astimezone(pytz.UTC)

    # Query schedules in UTC
    schedules = Schedule.query.filter(
        Schedule.technician_id == current_user.id,
        Schedule.start_time >= week_start_utc,
        Schedule.start_time < week_end_utc
    ).all()

    # Convert schedule times to user's timezone
    user_tz = current_user.get_timezone()
    for schedule in schedules:
        # Ensure times are timezone-aware in UTC
        if schedule.start_time.tzinfo is None:
            schedule.start_time = pytz.UTC.localize(schedule.start_time)
        if schedule.end_time.tzinfo is None:
            schedule.end_time = pytz.UTC.localize(schedule.end_time)

        # Convert to user's timezone
        schedule.start_time = schedule.start_time.astimezone(user_tz)
        schedule.end_time = schedule.end_time.astimezone(user_tz)

    form = ScheduleForm()
    form.technician.choices = [(current_user.id, current_user.username)]
    form.technician.data = current_user.id
    
    # Set up location choices
    locations = Location.query.filter_by(active=True).order_by(Location.name).all()
    form.location_id.choices = [(l.id, l.name) for l in locations]
    # Add an empty choice if no locations exist
    if not locations:
        form.location_id.choices = [(0, 'No locations available')]

    # Debug mobile detection
    print(f"is_mobile_device() in personal_schedule: {is_mobile_device()}")
    is_mobile = is_mobile_device()  # Force evaluation
    print(f"is_mobile value in personal_schedule: {is_mobile}")
    
    # Check if user is on a mobile device
    if is_mobile_device():
        print("Using mobile template for personal schedule")
        # Get active tickets for the sidebar
        from ticket_routes import get_active_sidebar_tickets
        active_sidebar_tickets = get_active_sidebar_tickets()
        
        return render_template('mobile_personal_schedule.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            today=datetime.now(current_user.get_timezone()),
                            datetime=datetime,
                            timedelta=timedelta,
                            personal_view=True,
                            active_sidebar_tickets=active_sidebar_tickets)
    else:
        return render_template('personal_schedule.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            today=datetime.now(current_user.get_timezone()),
                            datetime=datetime,
                            timedelta=timedelta,
                            personal_view=True)

@app.route('/admin/export_schedules')
@login_required
def export_schedules():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            flash('Please select both start and end dates.')
            return redirect(url_for('admin_dashboard'))

        # Convert dates to UTC datetime objects
        start_datetime = pytz.timezone('America/Chicago').localize(
            datetime.strptime(start_date, '%Y-%m-%d')
        ).astimezone(pytz.UTC)
        end_datetime = (pytz.timezone('America/Chicago').localize(
            datetime.strptime(end_date, '%Y-%m-%d')
        ) + timedelta(days=1)).astimezone(pytz.UTC)

        # Create a new Excel workbook
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        # Get all users sorted by username
        users = User.query.order_by(User.username).all()

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        for user in users:
            # Get user's schedules for the date range
            schedules = Schedule.query.filter(
                Schedule.technician_id == user.id,
                Schedule.start_time >= start_datetime,
                Schedule.start_time < end_datetime
            ).order_by(Schedule.start_time).all()

            if not schedules:
                continue

            # Create a new worksheet for each user
            ws = wb.create_sheet(title=user.username[:31])  # Excel limits sheet names to 31 chars

            # Write header
            ws['A1'] = f'Schedule Export - {user.username}'
            ws['A2'] = f'Period: {start_date} to {end_date}'

            # Calculate total hours
            user_tz = pytz.timezone('America/Chicago')
            total_minutes = 0

            for schedule in schedules:
                start_time = schedule.start_time.astimezone(user_tz)
                end_time = schedule.end_time.astimezone(user_tz)
                duration = (end_time - start_time).total_seconds() / 60
                total_minutes += duration

            total_hours = total_minutes // 60
            ws['A4'] = 'Total Hours:'
            ws['B4'] = f"{total_hours:.0f}:00:00"

            # Write column headers
            headers = ['Day', 'Date', 'Clock In', 'Clock Out', 'Total', 'Type', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill

            row = 7
            date_cursor = start_datetime.astimezone(user_tz)
            end_date = end_datetime.astimezone(user_tz)

            while date_cursor.date() < end_date.date():
                day_schedules = [s for s in schedules 
                               if s.start_time.astimezone(user_tz).date() == date_cursor.date()]

                if day_schedules:
                    for schedule in sorted(day_schedules, key=lambda s: s.start_time):
                        start_time = schedule.start_time.astimezone(user_tz)
                        end_time = schedule.end_time.astimezone(user_tz)
                        minutes = int((end_time - start_time).total_seconds() / 60)
                        hours = minutes // 60

                        # Determine the type and notes
                        entry_type = "Time Off" if schedule.time_off else "Work"
                        notes = []
                        if schedule.location:
                            notes.append(f"Location: {schedule.location.name}")
                        if schedule.time_off:
                            notes.append("TIME OFF")
                        if schedule.description:
                            if "ON-CALL" in schedule.description.upper():
                                notes.append("ON-CALL")
                            elif "PLEX" in schedule.description.upper():
                                notes.append("PLEX")
                            else:
                                notes.append(schedule.description)

                        # Write schedule data
                        ws.cell(row=row, column=1).value = start_time.strftime('%A')
                        ws.cell(row=row, column=2).value = start_time.strftime('%-m/%-d/%Y')
                        ws.cell(row=row, column=3).value = start_time.strftime('%-I:%M %p')
                        ws.cell(row=row, column=4).value = end_time.strftime('%-I:%M %p')
                        ws.cell(row=row, column=5).value = f"{hours}:00"
                        ws.cell(row=row, column=6).value = entry_type
                        ws.cell(row=row,column=7).value = " | ".join(notes) if notes else ""
                        row += 1
                else:
                    # Write empty row for days with no schedule
                    ws.cell(row=row, column=1).value= date_cursor.strftime('%A')
                    ws.cell(row=row, column=2).value = date_cursor.strftime('%-m/%-d/%Y')
                    ws.cell(row=row, column=3).value = "0"
                    ws.cell(row=row, column=4).value = "0"
                    ws.cell(row=row, column=5).value = "0:00"
                    row += 1

                date_cursor += timedelta(days=1)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'timesheets_{start_date}_to_{end_date}.xlsx'
        )

    except Exception as e:
        app.logger.error(f"Error exporting schedules: {str(e)}")
        flash('Error exporting schedules. Please try again.')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/quick_links')
@login_required
def admin_quick_links():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    form = QuickLinkForm()
    quick_links = QuickLink.query.order_by(QuickLink.order.asc()).all()
    return render_template('admin/quick_links.html', quick_links=quick_links, form=form)

@app.route('/admin/quick_links/create', methods=['POST'])
@login_required
def admin_create_quick_link():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        # Create new quick link
        link = QuickLink(
            title=request.form.get('title'),
            url=request.form.get('url'),
            icon=request.form.get('icon', 'link'),
            category=request.form.get('category'),
            order=request.form.get('order', 0)
        )

        db.session.add(link)
        db.session.commit()
        flash('Quick link created successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error creating quick link: {str(e)}")
        flash('Error creating quick link. Please try again.')

    return redirect(url_for('admin_quick_links'))

@app.route('/admin/quick_links/edit/<int:link_id>', methods=['POST'])
@login_required
def admin_edit_quick_link(link_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    link = QuickLink.query.get_or_404(link_id)
    try:
        link.title = request.form.get('title')
        link.url = request.form.get('url')
        link.icon = request.form.get('icon')
        link.category = request.form.get('category')
        link.order = request.form.get('order', 0)

        db.session.commit()
        flash('Quick link updated successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating quick link: {str(e)}")
        flash('Error updating quick link. Please try again.')

    return redirect(url_for('admin_quick_links'))

@app.route('/admin/quick_links/delete/<int:link_id>')
@login_required
def admin_delete_quick_link(link_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    link = QuickLink.query.get_or_404(link_id)
    try:
        db.session.delete(link)
        db.session.commit()
        flash('Quick link deleted successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting quick link: {str(e)}")
        flash('Error deleting quick link.')

    return redirect(url_for('admin_quick_links'))

@app.route('/admin/quick_links/reorder', methods=['POST'])
@login_required
def admin_reorder_quick_links():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403

    try:
        new_order = request.json
        # First, get all links and set a high order number to avoid conflicts
        links = QuickLink.query.all()
        for link in links:
            link.order = 10000 + link.order

        db.session.commit()

        # Then update with new order
        for item in new_order:
            link = QuickLink.query.get(item['id'])
            if link:
                link.order = int(item['order'])

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error reordering quick links: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

def get_open_tickets(limit=5):
    """Get open tickets for the current user (assigned to or created by)"""
    from models import Ticket, TicketStatus
    from flask_login import current_user
    from app import app
    
    if not current_user.is_authenticated:
        return []
    
    app.logger.debug(f"Fetching up to {limit} active tickets for sidebar display")
    
    # Get all active tickets (open, in progress, pending)
    # This is SEPARATE from the main ticket dashboard filtering
    query = Ticket.query.filter(
        Ticket.status.in_([TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.PENDING])
    ).order_by(
        # Order by priority (highest first) and then creation date (newest first)
        Ticket.priority.desc(),
        Ticket.created_at.desc()
    ).limit(limit)
    
    tickets = query.all()
    app.logger.debug(f"Found {len(tickets)} active tickets for sidebar display")
    
    return tickets

@app.context_processor
def inject_quick_links():
    def get_quick_links():
        return QuickLink.query.order_by(QuickLink.order.asc(), QuickLink.category).all()
    
    def get_user_tickets():
        # Always show active tickets in the sidebar regardless of main content filters
        return get_open_tickets(5)  # Limit to 5 tickets
    
    def get_active_sidebar_tickets():
        """
        This function always returns active tickets for the sidebar
        regardless of any filtering applied in the main dashboard.
        """
        from models import Ticket, TicketStatus
        from app import app
        
        app.logger.debug("Getting active tickets for sidebar (independent of dashboard filters)")
        
        # Get all active tickets (open, in_progress, pending) for sidebar
        query = Ticket.query.filter(
            Ticket.status.in_([TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.PENDING])
        ).order_by(
            # Order by priority (highest first) and then creation date (newest first)
            Ticket.priority.desc(),
            Ticket.created_at.desc()
        ).limit(5)
        
        tickets = query.all()
        app.logger.debug(f"Found {len(tickets)} active tickets for sidebar")
        
        return tickets
    
    return dict(
        get_quick_links=get_quick_links,
        get_user_tickets=get_user_tickets,
        get_active_sidebar_tickets=get_active_sidebar_tickets
    )

@app.route('/api/upcoming_time_off')
@login_required
def get_upcoming_time_off(for_template=False):
    """
    Get time off entries for the next 2 weeks
    If for_template is True, returns data formatted for template rendering
    Otherwise returns JSON response for API endpoint
    """
    try:
        # Get current time in UTC since our database stores times in UTC
        current_time = datetime.now(pytz.UTC)
        two_weeks_later = current_time + timedelta(days=14)

        # Query for upcoming time off entries
        time_off_entries = (Schedule.query
            .join(User)
            .filter(
                Schedule.start_time >= current_time,
                Schedule.start_time <= two_weeks_later,
                Schedule.time_off == True
            )
            .order_by(Schedule.start_time)
            .all())

        # Format data for template rendering with user color
        if for_template:
            template_entries = []
            for entry in time_off_entries:
                user = User.query.get(entry.technician_id)
                if user:
                    # Convert to user's timezone
                    user_tz = current_user.get_timezone()
                    start_time = entry.start_time.astimezone(user_tz)
                    end_time = entry.end_time.astimezone(user_tz)
                    
                    template_entries.append({
                        'username': user.username,
                        'color': user.color,
                        'start_time': start_time.strftime('%b %d, %I:%M %p'),
                        'end_time': end_time.strftime('%b %d, %I:%M %p'),
                        'description': entry.description
                    })
            return template_entries
        
        # Original API response formatting
        time_off_data = []
        user_tz = pytz.timezone('America/Los_Angeles')  # Default timezone
        
        # Group entries by username and consolidate consecutive dates
        user_entries = {}
        formatted_entries = []

        for entry in time_off_entries:
            user = User.query.get(entry.technician_id)
            if not user:
                continue
                
            username = user.username
            
            if username not in user_entries:
                user_entries[username] = []

            start_local = entry.start_time.astimezone(user_tz)
            end_local = entry.end_time.astimezone(user_tz)

            user_entries[username].append({
                'start_date': start_local.date(),
                'end_date': end_local.date(),
                'description': entry.description,
                'color': user.color
            })

        # Consolidate consecutive dates for each user
        for username, entries in user_entries.items():
            if not entries:
                continue
                
            entries.sort(key=lambda x: x['start_date'])
            consolidated = []
            current_entry = entries[0]

            for entry in entries[1:]:
                if (entry['start_date'] - current_entry['end_date']).days <= 1:
                    # Consecutive days, extend the current entry
                    current_entry['end_date'] = max(current_entry['end_date'], entry['end_date'])
                else:
                    # Non-consecutive, add current entry and start a new one
                    consolidated.append(current_entry)
                    current_entry = entry

            consolidated.append(current_entry)

            # Format consolidated entries
            for entry in consolidated:
                duration = (entry['end_date'] - entry['start_date']).days + 1
                formatted_entries.append({
                    'username': username,
                    'start_date': entry['start_date'].strftime('%b %d'),
                    'end_date': entry['end_date'].strftime('%b %d'),
                    'duration': f"{duration} day{'s' if duration != 1 else ''}",
                    'description': entry.get('description') or 'Time Off',
                    'color': entry.get('color', '#3498db')
                })

        return jsonify(formatted_entries)
    except Exception as e:
        app.logger.error(f"Error in get_upcoming_time_off: {str(e)}")
        if for_template:
            return []
        return jsonify([])

@app.route('/admin/backup')
@login_required
def admin_backup():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    return render_template('admin/backup.html')

@app.route('/admin/backup/download')
@login_required
def download_backup():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        # Collect all data with complete reference information
        backup_data = {
            'users': [user.to_dict() for user in User.query.all()],
            'locations': [location.to_dict() for location in Location.query.all()],
            'schedules': [schedule.to_dict() for schedule in Schedule.query.all()],
            'quick_links': [link.to_dict() for link in QuickLink.query.all()],
            'ticket_categories': [category.to_dict() for category in TicketCategory.query.all()],
            'tickets': [ticket.to_dict() for ticket in Ticket.query.all()],  # Include all tickets (archived and non-archived)
            'email_settings': [settings.to_dict() for settings in EmailSettings.query.all()]
        }

        # Create the backup file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_json = json.dumps(backup_data, indent=2, default=str)

        response = make_response(backup_json)
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename=backup_{timestamp}.json'

        app.logger.info(f"Backup created successfully with {len(backup_data['schedules'])} schedules")
        return response

    except Exception as e:
        app.logger.error(f"Error creating backup: {str(e)}")
        flash('Error creating backup')
        return redirect(url_for('admin_backup'))

@app.route('/admin/restore', methods=['POST'])
@login_required
def restore_backup():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    if 'backup_file' not in request.files:
        flash('No file uploaded')
        return redirect(url_for('admin_backup'))

    file = request.files['backup_file']
    if not file.filename:
        flash('No file selected')
        return redirect(url_for('admin_backup'))

    try:
        backup_data = json.loads(file.read().decode('utf-8'))
        app.logger.info("Starting backup restoration process...")

        # Initialize mappings for existing data
        existing_users = {user.username.lower(): user for user in User.query.all()}
        existing_locations = {loc.name.lower(): loc for loc in Location.query.all()}
        existing_quick_links = {(link.title.lower(), link.url.lower()): link for link in QuickLink.query.all()}

        try:
            # Process users first
            if 'users' in backup_data:
                app.logger.info("Starting user restoration...")
                for user_data in backup_data['users']:
                    try:
                        username = user_data.get('username')
                        email = user_data.get('email')

                        if not username or not email:
                            app.logger.warning(f"Skipping user with missing data")
                            continue

                        # Check if user exists (case-insensitive)
                        existing_user = existing_users.get(username.lower())
                        if existing_user:
                            app.logger.info(f"User {username} already exists")
                            continue

                        # Create new user
                        user = User(
                            username=username,
                            email=email,
                            password_hash=user_data['password_hash'],
                            color=user_data.get('color', '#3498db'),
                            is_admin=user_data.get('is_admin', False),
                            timezone=user_data.get('timezone', 'America/Los_Angeles')
                        )
                        db.session.add(user)
                        app.logger.info(f"Created new user {username}")

                    except Exception as e:
                        app.logger.error(f"Error processing user {username}: {str(e)}")
                        continue

                db.session.commit()
                app.logger.info("Users committed successfully")

            # Process locations
            if 'locations' in backup_data:
                app.logger.info("Starting location restoration...")
                for loc_data in backup_data['locations']:
                    try:
                        name = loc_data.get('name')
                        if not name:
                            continue

                        # Check if location exists (case-insensitive)
                        existing_location = existing_locations.get(name.lower())
                        if existing_location:
                            app.logger.info(f"Location {name} already exists")
                            continue

                        # Create new location
                        location = Location(
                            name=name,
                            description=loc_data.get('description', ''),
                            active=loc_data.get('active', True)
                        )
                        db.session.add(location)
                        app.logger.info(f"Created new location {name}")

                    except Exception as e:
                        app.logger.error(f"Error processing location {name}: {str(e)}")
                        continue

                db.session.commit()
                app.logger.info("Locations committed successfully")

            # Process schedules
            restored_count = 0
            skipped_count = 0

            if 'schedules' in backup_data:
                app.logger.info("Starting schedule restoration...")

                # Refresh user and location data after commits
                users_by_username = {user.username: user for user in User.query.all()}
                locations_by_name = {loc.name: loc for loc in Location.query.all()}

                for schedule_data in backup_data['schedules']:
                    try:
                        # Get technician by username
                        tech_username = schedule_data.get('technician_username')
                        if not tech_username:
                            app.logger.warning("Schedule missing technician username")
                            skipped_count += 1
                            continue

                        technician = users_by_username.get(tech_username)
                        if not technician:
                            app.logger.warning(f"Could not find technician: {tech_username}")
                            skipped_count += 1
                            continue

                        # Get location by name if specified
                        location_id = None
                        location_name = schedule_data.get('location_name')
                        if location_name:
                            location = locations_by_name.get(location_name)
                            if location:
                                location_id = location.id

                        start_time = datetime.fromisoformat(schedule_data['start_time'])
                        end_time = datetime.fromisoformat(schedule_data['end_time'])

                        # Check for existing schedule with same technician, time, and location
                        existing_schedule = Schedule.query.filter_by(
                            technician_id=technician.id,
                            start_time=start_time,
                            end_time=end_time,
                            location_id=location_id
                        ).first()

                        if existing_schedule:
                            app.logger.info(f"Schedule already exists for {tech_username} at {start_time}")
                            skipped_count += 1
                            continue

                        schedule = Schedule(
                            technician_id=technician.id,
                            location_id=location_id,
                            start_time=start_time,
                            end_time=end_time,
                            description=schedule_data.get('description'),
                            time_off=schedule_data.get('time_off', False)
                        )
                        db.session.add(schedule)
                        restored_count += 1

                    except Exception as e:
                        app.logger.error(f"Error processing schedule: {str(e)}")
                        skipped_count += 1
                        continue

                db.session.commit()
                app.logger.info(f"Schedules committed successfully. Restored: {restored_count}, Skipped: {skipped_count}")

            # Process quick links if present
            if 'quick_links' in backup_data:
                app.logger.info("Starting quick links restoration...")
                for link_data in backup_data['quick_links']:
                    try:
                        title = link_data.get('title', '').lower()
                        url = link_data.get('url', '').lower()

                        # Skip if title or url is missing
                        if not title or not url:
                            continue

                        # Check if link already exists
                        if (title, url) in existing_quick_links:
                            app.logger.info(f"Quick link already exists: {title}")
                            continue

                        link = QuickLink(
                            title=link_data['title'],
                            url=link_data['url'],
                            icon=link_data.get('icon', 'link'),
                            category=link_data['category'],
                            order=link_data.get('order', 0)
                        )
                        db.session.add(link)
                        app.logger.info(f"Created new quick link: {title}")

                    except Exception as e:
                        app.logger.error(f"Error processing quick link: {str(e)}")
                        continue

                db.session.commit()
                app.logger.info("Quick links committed successfully")
                
            # Process ticket categories if present
            if 'ticket_categories' in backup_data:
                app.logger.info("Starting ticket categories restoration...")
                existing_categories = {cat.name.lower(): cat for cat in TicketCategory.query.all()}
                
                for category_data in backup_data['ticket_categories']:
                    try:
                        name = category_data.get('name')
                        if not name:
                            continue
                            
                        # Check if category already exists
                        if name.lower() in existing_categories:
                            app.logger.info(f"Ticket category {name} already exists")
                            continue
                            
                        # Create new category
                        category = TicketCategory(
                            name=name,
                            description=category_data.get('description', ''),
                            icon=category_data.get('icon', 'help-circle'),
                            priority_level=category_data.get('priority_level', 0)
                        )
                        db.session.add(category)
                        app.logger.info(f"Created new ticket category: {name}")
                        
                    except Exception as e:
                        app.logger.error(f"Error processing ticket category: {str(e)}")
                        continue
                        
                db.session.commit()
                app.logger.info("Ticket categories committed successfully")
                
            # Process tickets if present
            tickets_restored = 0
            tickets_skipped = 0
            
            if 'tickets' in backup_data:
                app.logger.info("Starting tickets restoration...")
                
                # Refresh reference data
                users_by_username = {user.username: user for user in User.query.all()}
                categories_by_name = {cat.name: cat for cat in TicketCategory.query.all()}
                
                for ticket_data in backup_data['tickets']:
                    try:
                        title = ticket_data.get('title')
                        description = ticket_data.get('description')
                        
                        if not title or not description:
                            app.logger.warning("Ticket missing title or description")
                            tickets_skipped += 1
                            continue
                            
                        # Find category by name
                        category_name = ticket_data.get('category_name')
                        if not category_name or category_name not in categories_by_name:
                            app.logger.warning(f"Could not find category: {category_name}")
                            tickets_skipped += 1
                            continue
                            
                        # Find creator by username
                        creator_username = ticket_data.get('creator_username')
                        if not creator_username or creator_username not in users_by_username:
                            app.logger.warning(f"Could not find creator: {creator_username}")
                            tickets_skipped += 1
                            continue
                            
                        # Find assigned user by username if present
                        assigned_to = None
                        assigned_username = ticket_data.get('assigned_username')
                        if assigned_username and assigned_username in users_by_username:
                            assigned_to = users_by_username[assigned_username].id
                            
                        # Parse dates
                        created_at = None
                        if ticket_data.get('created_at'):
                            created_at = datetime.fromisoformat(ticket_data['created_at'])
                            
                        updated_at = None
                        if ticket_data.get('updated_at'):
                            updated_at = datetime.fromisoformat(ticket_data['updated_at'])
                            
                        due_date = None
                        if ticket_data.get('due_date'):
                            due_date = datetime.fromisoformat(ticket_data['due_date'])
                            
                        # Check if ticket with the original ID already exists
                        ticket_id = ticket_data.get('id')
                        existing_ticket = None
                        if ticket_id:
                            existing_ticket = Ticket.query.get(ticket_id)
                            
                        if existing_ticket:
                            app.logger.info(f"Ticket ID {ticket_id} already exists, skipping")
                            tickets_skipped += 1
                            continue
                            
                        # Create the ticket with its original ID to properly restore it
                        from sqlalchemy import text
                        
                        # First create the ticket without ID
                        ticket = Ticket(
                            title=title,
                            description=description,
                            category_id=categories_by_name[category_name].id,
                            status=ticket_data.get('status', TicketStatus.OPEN),
                            priority=ticket_data.get('priority', 0),
                            assigned_to=assigned_to,
                            created_by=users_by_username[creator_username].id,
                            created_at=created_at,
                            updated_at=updated_at,
                            due_date=due_date,
                            archived=ticket_data.get('archived', False)
                        )
                        
                        # Set the ID to match the original if provided
                        if ticket_id:
                            # Use raw SQL to set the ID explicitly
                            db.session.execute(text("ALTER SEQUENCE ticket_id_seq RESTART WITH :next_id"), 
                                            {"next_id": int(ticket_id) + 1})
                            ticket.id = ticket_id
                            
                        db.session.add(ticket)
                        db.session.flush()  # Get the ticket ID before adding comments
                        
                        # Process comments if present
                        if 'comments' in ticket_data:
                            for comment_data in ticket_data['comments']:
                                user_username = comment_data.get('username')
                                if not user_username or user_username not in users_by_username:
                                    # Try to use system user for comments if the original user isn't found
                                    system_user = User.query.filter_by(username="System").first()
                                    if not system_user:
                                        # Create system user if needed
                                        system_user = User(
                                            username="System", 
                                            email="system@example.com",
                                            is_admin=False
                                        )
                                        system_user.set_password(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20)))
                                        db.session.add(system_user)
                                        db.session.flush()
                                    user_id = system_user.id
                                else:
                                    user_id = users_by_username[user_username].id
                                    
                                # Get original comment ID if available
                                comment_id = comment_data.get('id')
                                
                                # Create the comment
                                comment = TicketComment(
                                    ticket_id=ticket.id,
                                    user_id=user_id,
                                    content=comment_data.get('content', ''),
                                    created_at=datetime.fromisoformat(comment_data['created_at']) if comment_data.get('created_at') else None,
                                    updated_at=datetime.fromisoformat(comment_data['updated_at']) if comment_data.get('updated_at') else None
                                )
                                
                                # Restore the original ID if possible
                                if comment_id:
                                    db.session.execute(text("ALTER SEQUENCE ticket_comment_id_seq RESTART WITH :next_id"), 
                                                    {"next_id": int(comment_id) + 1})
                                    comment.id = comment_id
                                    
                                db.session.add(comment)
                                
                        # Process history entries if present
                        if 'history' in ticket_data:
                            for history_data in ticket_data['history']:
                                user_username = history_data.get('username')
                                if not user_username or user_username not in users_by_username:
                                    # Try to use system user for history entries if the original user isn't found
                                    system_user = User.query.filter_by(username="System").first()
                                    if not system_user:
                                        # Create system user if needed
                                        system_user = User(
                                            username="System", 
                                            email="system@example.com",
                                            is_admin=False
                                        )
                                        system_user.set_password(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20)))
                                        db.session.add(system_user)
                                        db.session.flush()
                                    user_id = system_user.id
                                else:
                                    user_id = users_by_username[user_username].id
                                
                                # Get original history ID if available
                                history_id = history_data.get('id')
                                
                                # Create the history entry
                                history = TicketHistory(
                                    ticket_id=ticket.id,
                                    user_id=user_id,
                                    action=history_data.get('action', ''),
                                    details=history_data.get('details', ''),
                                    created_at=datetime.fromisoformat(history_data['created_at']) if history_data.get('created_at') else None
                                )
                                
                                # Restore the original ID if possible
                                if history_id:
                                    db.session.execute(text("ALTER SEQUENCE ticket_history_id_seq RESTART WITH :next_id"), 
                                                    {"next_id": int(history_id) + 1})
                                    history.id = history_id
                                    
                                db.session.add(history)
                                
                        tickets_restored += 1
                        
                    except Exception as e:
                        app.logger.error(f"Error processing ticket: {str(e)}")
                        tickets_skipped += 1
                        continue
                        
                db.session.commit()
                app.logger.info(f"Tickets committed successfully. Restored: {tickets_restored}, Skipped: {tickets_skipped}")
                
            # Process email settings if present
            if 'email_settings' in backup_data and backup_data['email_settings']:
                app.logger.info("Restoring email settings...")
                try:
                    settings_data = backup_data['email_settings'][0]
                    settings = EmailSettings.query.first()
                    
                    if not settings:
                        settings = EmailSettings()
                        db.session.add(settings)
                        
                    settings.admin_email_group = settings_data.get('admin_email_group', 'alerts@obedtv.com')
                    settings.notify_on_create = settings_data.get('notify_on_create', True)
                    settings.notify_on_update = settings_data.get('notify_on_update', True)
                    settings.notify_on_delete = settings_data.get('notify_on_delete', True)
                    
                    db.session.commit()
                    app.logger.info("Email settings restored successfully")
                except Exception as e:
                    app.logger.error(f"Error restoring email settings: {str(e)}")

            flash(f'Backup restored successfully! {restored_count} schedules restored, {skipped_count} skipped. {tickets_restored} tickets restored, {tickets_skipped} skipped.')
            app.logger.info("Backup restore completed successfully")

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error in backup restoration: {str(e)}")
            flash('Error restoring backup')
            raise

    except json.JSONDecodeError:
        flash('Invalid backup file format')
    except Exception as e:
        flash('Error restoring backup')
        app.logger.error(f"Unexpected error in restore_backup: {str(e)}")

    return redirect(url_for('admin_backup'))

@app.route('/admin/locations/edit/<int:location_id>', methods=['POST'])
@login_required
def admin_edit_location(location_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    location = Location.query.get_or_404(location_id)
    try:
        # Get form data with proper boolean conversion for active status
        name = request.form.get('name')
        description = request.form.get('description', '')
        # Explicitly check for the checkbox value
        active = 'active' in request.form

        # Validate required fields
        if not name:
            flash('Location name is required.')
            return redirect(url_for('admin_locations'))

        # Track if status changed
        status_changed = location.active != active

        # Update location
        location.name = name
        location.description = description
        location.active = active

        db.session.commit()

        # Log the status change if it occurred
        if status_changed:
            app.logger.info(f"Location '{location.name}' status changed to {'active' if active else 'inactive'}")
            flash(f"Location '{location.name}' has been {'activated' if active else 'deactivated'}.")
        else:
            flash('Location updated successfully!')

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating location: {str(e)}")
        flash('Error updating location. Please try again.')

    return redirect(url_for('admin_locations'))

@app.route('/admin/locations/delete/<int:location_id>')
@login_required
def admin_delete_location(location_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    location = Location.query.get_or_404(location_id)

    # Check if location is being used in any schedules
    if location.schedules:
        flash('Cannot delete location that has associated schedules.')
        return redirect(url_for('admin_locations'))

    try:
        db.session.delete(location)
        db.session.commit()
        flash('Location deleted successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting location: {str(e)}")
        flash('Error deleting location. Please try again.')

    return redirect(url_for('admin_locations'))
@app.route('/admin/email-settings', methods=['GET', 'POST'])
@login_required
def admin_email_settings():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    settings = EmailSettings.query.first()
    if not settings:
        settings = EmailSettings()
        db.session.add(settings)
        db.session.commit()

    form = EmailSettingsForm(obj=settings)

    if form.validate_on_submit():
        try:
            settings.admin_email_group = form.admin_email_group.data
            settings.notify_on_create = form.notify_on_create.data
            settings.notify_on_update = form.notify_on_update.data
            settings.notify_on_delete = form.notify_on_delete.data
            db.session.commit()
            flash('Email settings updated successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating email settings: {str(e)}")
            flash('Error updating email settings.')

    return render_template('admin/email_settings.html', form=form)